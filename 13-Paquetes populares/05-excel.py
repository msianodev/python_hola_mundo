from openpyxl import load_workbook

# Cargar el archivo Excel existente
archivo_entrada = "empleados.xlsx"
archivo_salida = "empleados_modificado.xlsx"

wb = load_workbook(archivo_entrada)
ws = wb.active  # Seleccionar la hoja activa

# Eliminar la fila 5 (puedes cambiar el número según sea necesario)
ws.delete_rows(5)

# Modificar el correo del empleado en la fila 3 (columna D -> "Correo")
ws["D3"] = "nuevo.email@empresa.com"

# Guardar los cambios en un nuevo archivo
wb.save(archivo_salida)

print(f"Archivo guardado como {archivo_salida}")
